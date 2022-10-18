from django.shortcuts import render

# Create your views here.
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Group
from django.contrib import messages
from django.conf import settings

from django.views.decorators.csrf import csrf_protect
from django.shortcuts import render, redirect, get_object_or_404
from django.template import RequestContext
from django.template.loader import render_to_string
from django.http import HttpResponse, HttpResponseRedirect
from django.core.mail import BadHeaderError, send_mail, EmailMessage, EmailMultiAlternatives

from django.utils.html import strip_tags

#from random_username.generate import generate_username
#from random_password import random_password
from .forms import SignUpForm, CustomerForm, OrderForm, ConfirmedOrderForm, ProductsSheetForm, ChiqueForm, SalesOrderForm, BillForm
from .decorators import *
from datetime import datetime
from .models import *
import xlrd, os, nexmo
from io import BytesIO 
from openpyxl import load_workbook
from django.utils.translation import gettext_lazy as _

####################################################################################
def index(request):
    return render(request, "bayanApp/home.html")
#####################################################################################
@unauthenticated_user
def signUp(request):
    form = SignUpForm()
    if request.method == 'POST':
        form = SignUpForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            group = Group.objects.get(name=_('vendors'))
            user.is_active = False
            user.groups.add(group)
            user.save()
#            Vendor.objects.create(user=user, VEND_DESC_ARABIC=request.POST.get('VEND_DESC_ARABIC'))
            return redirect('login')
        return render(request, 'signup.html', {'form':form})
    context={'form': form}
    return render(request, 'signup.html', context)
####################################################################################
@unauthenticated_user
#@csrf_protect
#@ensure_csrf_cookie
def loginPage(request):
    if request.method =="POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            try:
                user = User.objects.get(username=username)
                if not user.is_active:
                    messages.info(request, _("User is not active"))
                else:            
                    messages.info(request, _("Username OR Password is incorrect"))
            except User.DoesNotExist:
                messages.info(request, _("Account does not exist"))
    context={}
    return render(request, 'login.html', context)
####################################################################################
@login_required(login_url='login')
def home(request):
    vendor = items = {}
    user = request.user
    p_type=_("Vendor")
    if user.person_type == p_type:
        try:
            vendor = Vendor.objects.get(user=user)
        except Vendor.DoesNotExist:
            return render(request, 'home.html', {})

        unread_transactions = Transaction.objects.filter(t_state = "unread", VEND_CODE=vendor.VEND_CODE)
        t_count = Transaction.objects.all().count()
        print(request.user.VEND_CODE)
        return render(request, 'home.html', {'un_transactions':unread_transactions, 't_count':t_count, "person":user.person_type, "p_type":p_type})
    if user.is_superuser :
        return redirect('/admin')
    return render(request, 'home.html', {"person":user.person_type, "p_type":p_type})

####################################################################################
@login_required(login_url='login')
def logoutUser(request):
    logout(request)
    return redirect('login')
####################################################################################
@login_required(login_url='login')
def profile(request):
    context = {}
    return render(request, 'profile.html', context)
####################################################################################
@login_required(login_url='login')
def updateProfile(request):
    user = request.user
    form = CustomerForm(instance=user)
    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            return redirect('profile')
    context = {'form':form}
    return render(request, "update_profile.html", context)
####################################################################################
@login_required(login_url='login')
def changePassword(request):
    if request.method =='POST':
        form = PasswordChangeForm(data=request.POST, user=request.user)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            return redirect('profile')
        else:
            return render(request, 'change_password.html', {'form':form})
    else:
        form = PasswordChangeForm(user=request.user)
        args = {'form':form}
        return render(request, 'change_password.html', args)
####################################################################################
def resetPassword(request):
    if request.method == "POST":
        return render(request, 'reset_password.html', {})
    return render(request, 'reset_password.html', args)
####################################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('vendors')])
def orderXo(request):
    form = OrderForm()
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
#            table_data = request.POST.get("table_data")
            print("quantity", form['quantity'].value())
#            myrequest = request.POST.copy()
#            form = OrderForm(myrequest)
            print("before form save")
#            data = eval(table_data)
            form.save()
            print("form saved")
            order = Order_MS_VCHR_XO.objects.get(order_no = request.POST.get("order_no"))
            vendor = Vendor.objects.get(VEND_CODE=order.VEND_CODE)
            order.vendor = vendor
            print("save order")
            order.save()
            return redirect('print_order/'+request.POST.get("order_no")+'/')
    order_no = datetime.now().strftime("%y%m%d%H%M%S")
    print(request.user.VEND_CODE)
    items = Product.objects.filter(VEND_CODE=request.user.VEND_CODE)
#    print(items[0].PACK_ID)
#    print(items[0].UNIT_DESC)
    context = {
            'form':form,
            'vendor':Vendor.objects.get(VEND_CODE=request.user.VEND_CODE),#request.user.vendor,
            'items':items,
            'order_no':order_no,
            'date':datetime.now().strftime("%y-%m-%d %H:%M:%S"),
            }
    return render(request, 'order_xo.html', context)
####################################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def salesOrder(request):
    form = SalesOrderForm()
    if request.method == 'POST':
        form = SalesOrderForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            order  = Order_MS_VCHR_XO.objects.get(order_no = form.cleaned_data["order_no"])
            vendor = Vendor.objects.get(VEND_DESC_ARABIC = form.cleaned_data["VEND_DESC_ARABIC"])
            order.VEND_CODE = vendor.VEND_CODE
            order.vendor = vendor
            order.order_image = order.image_data
            order.save()
            addTransaction("new order", "new order", order.VEND_CODE, order.order_no)
            return redirect('print_order/'+request.POST.get("order_no")+'/')
        context={'error':form.errors}
        return render(request, 'sales_order.html', context)
    context = {
            'form':form,
            'order_no'  :datetime.now().strftime("%y%m%d%H%M%S"),
            'date'      :datetime.now().strftime("%y-%m-%d %H:%M:%S"),
        }
    return render(request, 'sales_order.html', context)
######################################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def listAllOrders(request):
    orders={}
    if request.user.person_type == _("Sales"):
        orders = Order_MS_VCHR_XO.objects.all()
    return render(request, 'list_all_orders.html',{'orders':orders, 'mydate':datetime.now().strftime("%d-%B-%Y")})
#########################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('vendors'), _('sales')])
def listOrders(request, VEND_CODE):
    if request.user.person_type == _("Vendor") :
        if request.user.VEND_CODE != VEND_CODE:
            return redirect('my403')
    try:
        vendor = Vendor.objects.get(VEND_CODE=VEND_CODE)
        orders = Order_MS_VCHR_XO.objects.filter(VEND_CODE=VEND_CODE)
    except Vendor.DoesNotExist:
        return render(request, '404.html', {})
        
    context ={'orders':orders, 'vendor':vendor, 'mydate':datetime.now().strftime("%d-%B-%Y")} 
    return render(request, 'list_orders.html', context)
########################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales'),_('vendors')])
def printOrder(request, order_no):
    print(order_no)
    try:
        order = Order_MS_VCHR_XO.objects.get(order_no=order_no)
    except Order_MS_VCHR_XO.DoesNotExist:
        return render(request, '404.html', {})

    print(order.table_data)
    if order.table_data :
        order.table_data = eval(order.table_data)
    conf_order = None
    context = {'order':order, 'conf_order':conf_order, 'mydate':datetime.now().strftime("%d-%B-%Y")}
    return render(request, 'print_order.html', context)
#######################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def uploadConfOrder(request):
    if request.method == 'GET':
        form = ConfirmedOrderForm(request.GET)
        return HttpResponse(form)
    if request.method == 'POST':
        order =  Order_MS_VCHR_XO.objects.get(order_no=request.POST.get('order_no'))
        form =ConfirmedOrderForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            form.save() 
            order.order_state = _("Closed")
            order.save()
            to = [order.vendor.user.email]
            my_file = request.FILES['order_image']
            addTransaction("confirm order", "confirm order", order.VEND_CODE, order.order_no)
            sendEmail("Confirmed Order", to, order.order_no, my_file, "email.html")
            return redirect('print_order', order.order_no)
        else:
            return render(request, 'error_filetype.html', {'error':form.errors['order_image'], 'page':'print_orders', "parameter":order.order_no})

######################################################################
def sendEmail(subject, to, order_no, my_file, file_html):
    subject      = subject 
    from_email   = "admin@bayancoopq8.com"
    to           = to 
    html_content = render_to_string(file_html, {'order_no':order_no}) # render with dynamic value
    text_content = strip_tags(html_content) # Strip the html tag. So pCurvy wet plumber gags on masseurs cockeople can see the pure text at least.
#    email = EmailMessage(subject, html_content, from_email, to)
#    attached_file = os.path.join(settings.MEDIA_ROOT, my_file.name)
#    email.attach_file(attached_file)
    email      = EmailMultiAlternatives(subject, text_content, from_email, to)
    email.attach_alternative(html_content, "text/html")
    attached_file = os.path.join(settings.MEDIA_ROOT, my_file.name)
    email.attach_file(attached_file )
    email.send()

######################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('vendors')])
def deleteOrder(request, order_no):
    print('delete order')
    order = Order_MS_VCHR_XO.objects.get(order_no=order_no)
    if order.order_state == _("Opened"):
        print("opened")
        order.delete()
        return HttpResponse('The order was deleted')
    else:
        return HttpResponse("You can't delete this order")
#######################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def deleteOldOrders(request, day, month, year):
    print("delte old orders")
    import datetime
    orders = Order_MS_VCHR_XO.objects.filter(order_date__lte=datetime.date(int(year), int(month), int(day)))
    for order in orders:
        order.delete()
    return HttpResponse('The old orders was deleted')
#######################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales'), _('financials')])
def listVendors(request):
    vendors = Vendor.objects.all()
    return render(request, 'list_vendors.html', {'vendors':vendors, 'mydate':datetime.now().strftime("%d-%B-%Y")} )
######################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def listNewUsers(request):
    newUsers = list(User.objects.filter(is_active = False)) #.values_list('id', flat=True))
#    vendors = list(Vendor.objects.filter(VEND_CODE= none))
#    vendors = Vendor.objects.filter(user_id__in=list(User.objects.filter(is_active = False).values_list('id', flat=True)))
    return render(request, 'list_newUsers.html', {'vendors':newUsers, 'newUsers':newUsers, 'mydate':datetime.now().strftime("%d-%B-%Y")} )

######################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales'), _('financials')])
def listVendorInfo(request, VEND_CODE):
    vendor = Vendor.objects.get(VEND_CODE= VEND_CODE)
    return render(request, 'list_vendor_info.html', {"vendor":vendor} )

######################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def listNewUserInfo(request, user_id):
    user = User.objects.get(id = user_id)
    return render(request, 'list_newUser_info.html', {"user":user} )

#####################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('financials')])
def uploadChique(request):
    if request.method == 'GET':
        vend_code = request.GET.get('vendor_id')
        form = ChiqueForm(request.GET)
        return HttpResponse(form)
    if request.method == 'POST':
        vendor = Vendor.objects.get(VEND_CODE=request.POST.get('vendor_id'))
        form = ChiqueForm(request.POST, request.FILES)
        if form.is_valid():
            form.instance.vendor = vendor
            form.instance.VEND_CODE = vendor.VEND_CODE
            form.instance.c_date = datetime.now()
            form.save()
            print("upload chique")
#            to = [vendor.user.phone]
#            sendSMS(to)
#            addTransaction("upload chique", "upload chique", vendor.VEND_CODE, request.FILES['chique_image'])
            to = [vendor.user.email]
            sendEmail("Chique is ready", to, "", request.FILES['chique_image'], "chique_email.html")
            return redirect('list_vendors')
        else:
            return render(request, 'error_filetype.html', {'error':form.errors['chique_image'], 'page':'listVendors'})
        return HttpResponse(form)
#####################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('vendors')])
def uploadBill(request):
    if request.method == 'GET':
        form = BillForm(request.GET)
        return HttpResponse(form)
    if request.method == 'POST':
        vendor = Vendor.objects.get(user_id=request.user.id)
        form = BillForm(request.POST, request.FILES)
        if form.is_valid():
            form.instance.VEND_CODE = vendor.VEND_CODE
            form.instance.vendor = vendor
            form.instance.b_date = datetime.now()
            form.save()
        else:
            return render(request, 'error_filetype.html', {'error':form.errors['bill_file'], 'page':'home'})
        return redirect('home')
#############################################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('vendors'), _('sales')])
def listProducts(request, VEND_CODE):
    user = request.user
    context = {}
    if user.person_type == _("Vendor"):
        if user.VEND_CODE != VEND_CODE:
            return redirect('my403')
    try:
        vendor = Vendor.objects.get(VEND_CODE=VEND_CODE)
    except Vendor.DoesNotExist:
        return render(request, '404.html', {})
    items = Product.objects.filter(VEND_CODE=VEND_CODE)
    context = {'vendor':vendor, 'items':items, 'mydate':datetime.now().strftime("%d-%B-%Y")}
    return render(request, 'list_products.html', context)
###############################################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def updateVendorProducts(request):
    if request.method == 'GET':
        form = ProductsSheetForm(request.GET)
        vend_code = request.GET.get('VEND_CODE')
        return HttpResponse(form)
    
    if request.method == 'POST':
        form = ProductsSheetForm(request.POST, request.FILES)
        file_in_memory = request.FILES['products_sheet']
        vend_code = request.POST.get('VEND_CODE')
        if form.is_valid():
            wb = load_workbook(file_in_memory.file)
            ws = wb[wb.sheetnames[0]]
            excel_data = list()
            for row in ws.iter_rows(min_row=2):
                mydict = Product(
#                    ID          = row[0].value,
                    ARABIC_NAME = row[1].value,
                    VEND_CODE   = row[2].value,     PACK_ID     = row[3].value,
                    NO_OF_ITEMS = row[4].value,     UNIT_DESC   = row[5].value,
                    PURCH_PRICE = row[6].value,     SALE_PRICE  = row[7].value,
                    BARCODE     = row[8].value
                )
                excel_data.append(mydict)
            products = Product.objects.filter(VEND_CODE=vend_code)
            products.delete()
            Product.objects.bulk_create(excel_data, 10)
            addTransaction("update products", "update products", vend_code, vend_code)
            return redirect('list_products', VEND_CODE=vend_code)
        else:
            return render(request, 'error_filetype.html', {'error':form.errors['products_sheet'], 'page':'update_products'})

###########################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('sales')])
def updateAllProducts(request):
    import time
    if request.method == 'GET':
        form = ProductsSheetForm(request.GET)
        return HttpResponse(form['products_sheet'])

    if request.method == 'POST':
        start = time.time()
        form = ProductsSheetForm(request.POST, request.FILES)
        file_in_memory = request.FILES['products_sheet']
        if form.is_valid():
            print("form is valid")
            wb = load_workbook(file_in_memory)
            ws = wb[wb.sheetnames[0]]
            excel_data = list()
            start = time.time()
            for row in ws.iter_rows(min_row=2):
                mydict = Product(
                    ARABIC_NAME = row[1].value,
                    VEND_CODE   = row[2].value,     PACK_ID     = row[3].value,
                    NO_OF_ITEMS = row[4].value,     UNIT_DESC   = row[5].value,
                    PURCH_PRICE = row[6].value,     SALE_PRICE  = row[7].value,
                    BARCODE     = row[8].value
                    )
                excel_data.append(mydict)
                print(row)
            print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
            products = Product.objects.all()
            products.delete()
            print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
            Product.objects.bulk_create(excel_data,10)
            return redirect('list_vendors')
        else:
            return render(request, 'error_filetype.html', {'error':form.errors['products_sheet'], 'page':'update_products'})

#########################################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('vendors')])
def listTransactions(request, VEND_CODE):
    transactions = Transaction.objects.all()
    try:
        vendor = Vendor.objects.get(VEND_CODE=VEND_CODE)
        if VEND_CODE != request.user.VEND_CODE:
            return render(request, '403.html', {})
    except Vendor.DoesNotExist:
        return render(request, '404.html', {})
        messages.info(request, _("Account does not exist"))
    context = {'transactions':transactions, 'vendor':vendor, 'mydate':datetime.now().strftime("%d-%B-%Y")}
    return render(request, 'list_transactions.html', context)
#########################################################################
def addTransaction(t_type, details, vend_code, parameter):
    print(datetime.now().strftime("%y%m%d%H%M%S"))
    Transaction.objects.create(
            t_no        = datetime.now().strftime("%y%m%d%H%M%S"),
            t_date      = datetime.now(),
            t_type      = t_type,
            t_state     = 'unread',
            t_details   = details ,
            VEND_CODE   = vend_code,
            parameter   = parameter)
#########################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('vendors')])
def readTransaction(request, t_no):
    transaction = Transaction.objects.filter(t_no=t_no).update(t_state = "read")
    return redirect('home')

#########################################################################
@login_required(login_url='login')
@allowed_users(allowed_roles=[_('financials')])
def listBills(request, VEND_CODE):
    bills = Bill.objects.filter(VEND_CODE=VEND_CODE)
    vendor = Vendor.objects.get(VEND_CODE=VEND_CODE)
    context = {'bills':bills, 'vendor':vendor, 'mydate':datetime.now().strftime("%d-%B-%Y")}
    return render(request, 'list_bills.html', context)
#########################################################################
def my404(request):
    return render(request, '404.html', {})

def my403(request):
    return render(request, '403.html', {})
###########################################################################################################################











########################################################################
#@login_required(login_url='login')
#def vendorOrders(request, vendor_id):
#    orders = Order_MS_VCHR_XO.objects.filter(vendor_id=vendor_id)
#    return render(request, 'list_orders.html',{'orders':orders, 'vendor':orders[0].vendor} )
########################################################################
#@login_required(login_url='login')
#def orderDetails(request, id):
#    order = Order_MS_VCHR_XO.objects.get(id=id)
#    order.table_data = eval(order.table_data)[0]
#    context = {'order':order}
#    return render(request, 'order_details.html', context)


#def showBill(request, VEND_CODE):


##################################################################################################
#def sendSMS(to):
#    client = nexmo.Client(key='aac2e6b5', secret='hhB32cCwfALghFN5')
#    responseData = client.send_message({
#        'from': '+201009847025',
#        'to'  : to, #'+965 9997 8716', #order.vendor.user.phone,
#        'text': 'Dear vendor, we informed you that your chique is ready.',
#    })
#    if responseData["messages"][0]["status"] == "0":
#        print("Message sent successfully.")
#    else:
#        print(f"Message failed with error: {responseData['messages'][0]['error-text']}")
#    import os
#    from twilio.rest import Client
#    account_sid = "ACfd1cd141e4122c61367d07a2759a89c6"
#    auth_token = "78fc7f8d10586ad0d021d7d9214f3769"
#    client = Client(account_sid, auth_token)
#    sms = client.messages.create(
#            to = "+201009847025",
#            from_ ="+12103616234",
#            body="we informed you that your chique is ready",
#            )
#    print(sms.sid)
##################################################################################################

#########################################################################
#@login_required(login_url='login')
#def newVendor(request):
#    form = SignUpForm()
#    if request.method == 'POST':
#        form = SignUpForm(request.POST)
#        if form.is_valid():
#            user = form.save()
#            username = form.cleaned_data.get('username')
#            group = Group.objects.get(name='customer')
#            user.groups.add(group)
#            Vendor.objects.create(user=user,)
#            messages.success(request, 'Account was creatd for ' + username )
#            return redirect('login')

#    context={'form': form}
#    return render(request, 'new_vendor.html', context)
############################################################
#import time
#start = time.time()
#end = time.time()-start

##                dictionary = dict(zip(keys, row_data))



#        from django.db.models import Count
#        orders = Order_MS_VCHR_XO.objects.values('VEND_DESC_ARABIC', 'order_state', 'vendor').annotate(ocount=Count('vendor'))
#        return render(request, 'sales_lst_orders.html', {'orders':orders, 'vendors':vendors} )



#        form = ConfirmedOrderForm(initial={'order_no':order_no})
#        form.fields['order_no'].widget.attrs['readonly'] = True
